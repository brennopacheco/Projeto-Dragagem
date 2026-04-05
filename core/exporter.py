"""core/exporter.py — Exportacao Excel (openpyxl) e PDF (reportlab)."""

from __future__ import annotations

import io
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# ---------------------------------------------------------------------------
# Constantes de layout
# ---------------------------------------------------------------------------

# Colunas do Excel/PDF (19 colunas)
_HEADERS_TOP = [
    "Data", "Mare", "", "Hora", "", "Area", "KP", "", "Linhas", "",
    "Sistema\n(x)", "Obs", "Dist.\nFundo (m)", "Angulo", "", "Dir.\n(xx)",
    "Velocidade", "", "Consumo\nDiesel",
]
_HEADERS_SUB = [
    "", "Status", "Ampl.", "Inicio", "Final", "", "Inicio", "Final",
    "De", "Ate", "", "", "", "Na Tela", "No Nozzle", "", "P/ Frente",
    "P/ Re", "",
]

# Merges da linha de grupo (1-indexed, row 7)
# (col_start, col_end) para colspan > 1
_MERGES_TOP = [
    (2, 3),   # Mare
    (4, 5),   # Hora
    (7, 8),   # KP
    (9, 10),  # Linhas
    (14, 15), # Angulo
    (17, 18), # Velocidade
]

# Larguras das colunas (em caracteres)
_COL_WIDTHS = [
    12,   # Data
    6,    # Status
    7,    # Ampl.
    7,    # Inicio
    7,    # Final
    9,    # Area
    7,    # KP Inicio
    7,    # KP Final
    7,    # Linha De
    7,    # Linha Ate
    8,    # Sistema
    14,   # Obs
    10,   # Dist Fundo
    8,    # Ang Tela
    9,    # Ang Nozzle
    8,    # Dir
    8,    # Vel Frente
    8,    # Vel Re
    9,    # Consumo
]

# Cores Van Oord
_BLUE = "193C89"
_BLUE_LIGHT = "2550A5"
_ORANGE = "F58A18"
_BG_EVEN = "F6F8FC"
_WHITE = "FFFFFF"


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def exportar_excel(
    trechos: list[dict],
    config: dict,
    semana_inicio: date,
) -> io.BytesIO:
    """Gera arquivo Excel com a programacao da semana.

    Retorna BytesIO pronto para enviar como download.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Programacao"

    # Estilos
    header_font = Font(name="Segoe UI", bold=True, color=_WHITE, size=9)
    sub_font = Font(name="Segoe UI", bold=False, color=_WHITE, size=8)
    title_font = Font(name="Segoe UI", bold=True, color=_BLUE, size=14)
    info_font = Font(name="Segoe UI", size=9)
    info_bold = Font(name="Segoe UI", bold=True, size=9)
    data_font = Font(name="Segoe UI", size=9)
    en_font = Font(name="Segoe UI", bold=True, size=9, color=_BLUE)
    vz_font = Font(name="Segoe UI", bold=True, size=9, color=_ORANGE)

    fill_header = PatternFill("solid", fgColor=_BLUE)
    fill_sub = PatternFill("solid", fgColor=_BLUE_LIGHT)
    fill_even = PatternFill("solid", fgColor=_BG_EVEN)

    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Larguras
    for i, w in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Cabecalho ----
    # Linha 2: titulo
    ws.merge_cells("A2:S2")
    c = ws["A2"]
    c.value = "PROGRAMACAO DE DRAGAGEM"
    c.font = title_font
    c.alignment = Alignment(horizontal="center")

    # Linha 4: Empresa + Project
    ws["A4"] = config.get("empresa", "")
    ws["A4"].font = info_bold
    ws.merge_cells("A4:N4")
    ws["O4"] = "Project"
    ws["O4"].font = info_bold
    ws["P4"] = config.get("project", "")
    ws["P4"].font = info_font
    ws.merge_cells("P4:S4")

    # Linha 5: Vessel + Site
    ws["A5"] = f"Vessel: {config.get('vessel', '')}"
    ws["A5"].font = info_font
    ws.merge_cells("A5:N5")
    ws["O5"] = f"Site: {config.get('site', '')}"
    ws["O5"].font = info_font
    ws.merge_cells("O5:S5")

    # Linha 6: Data
    semana_fim = semana_inicio + timedelta(days=6)
    ws["A6"] = f"Data: {semana_inicio.strftime('%d/%m/%Y')} - {semana_fim.strftime('%d/%m/%Y')}"
    ws["A6"].font = info_font
    ws.merge_cells("A6:S6")

    # ---- Headers (linha 7 e 8) ----
    row_top = 7
    row_sub = 8

    # Linha 7 — grupo
    for i, h in enumerate(_HEADERS_TOP, 1):
        cell = ws.cell(row=row_top, column=i, value=h)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = center
        cell.border = thin_border

    # Merges da linha de grupo
    for cs, ce in _MERGES_TOP:
        ws.merge_cells(start_row=row_top, start_column=cs, end_row=row_top, end_column=ce)

    # Rowspan para colunas sem sub-header (merge linhas 7-8)
    for i in (1, 6, 11, 12, 13, 16, 19):
        ws.merge_cells(start_row=row_top, start_column=i, end_row=row_sub, end_column=i)

    # Linha 8 — sub-headers (pular colunas com rowspan)
    rowspan_cols = {1, 6, 11, 12, 13, 16, 19}
    for i, h in enumerate(_HEADERS_SUB, 1):
        if i in rowspan_cols:
            continue
        cell = ws.cell(row=row_sub, column=i, value=h)
        cell.font = sub_font
        cell.fill = fill_sub
        cell.alignment = center
        cell.border = thin_border

    # ---- Dados (a partir da linha 9) ----
    row = 9
    last_date = ""
    for idx, t in enumerate(trechos):
        show_date = t["data"] != last_date
        last_date = t["data"]

        values = [
            t["data"] if show_date else "",
            t["status"],
            t["amplitude"],
            t["inicio"],
            t["fim"],
            t.get("area") or "",
            t.get("kp_inicio") or "",
            t.get("kp_final") or "",
            t.get("linha_de") or "",
            t.get("linha_ate") or "",
            t.get("sistema_dragagem") or "",
            t.get("observacoes") or "",
            t.get("dist_fundo") or "",
            t.get("ang_tela") or "",
            t.get("ang_nozzle") or "",
            t.get("direcao") or "",
            t.get("vel_frente") or "",
            t.get("vel_re") or "",
            t.get("consumo_diesel") or "",
        ]

        is_even = idx % 2 == 0
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = data_font
            cell.alignment = center
            cell.border = thin_border
            if is_even:
                cell.fill = fill_even

        # Cor do status
        status_cell = ws.cell(row=row, column=2)
        status_cell.font = en_font if t["status"] == "EN" else vz_font

        # Data em negrito alinhado a esquerda
        if show_date:
            ws.cell(row=row, column=1).font = info_bold
            ws.cell(row=row, column=1).alignment = left

        row += 1

    # ---- Legenda ----
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=19)
    leg = ws.cell(row=row, column=1)
    leg.value = (
        "(x) AE = Agitador Externo; AI = Agitador Interno; COR = Corrente; "
        "CL = Curva Lateral; DS = Desagregador; EH = Erosion Head ou Mass Flow; "
        "JB = Jet Bar; JL = Jato Lateral; RD = Reduction"
    )
    leg.font = Font(name="Segoe UI", size=7, color="666666")
    leg.alignment = Alignment(wrap_text=True)

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=19)
    leg2 = ws.cell(row=row, column=1)
    leg2.value = "(xx) Direction of the bow"
    leg2.font = Font(name="Segoe UI", size=7, color="666666")

    # Salvar em BytesIO
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def exportar_pdf(
    trechos: list[dict],
    config: dict,
    semana_inicio: date,
) -> io.BytesIO:
    """Gera PDF landscape A4 com a programacao da semana.

    Retorna BytesIO pronto para enviar como download.
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    # Cores
    blue = colors.HexColor("#193C89")
    blue_light = colors.HexColor("#2550A5")
    orange = colors.HexColor("#F58A18")
    bg_even = colors.HexColor("#F6F8FC")
    white = colors.white

    # ---- Titulo ----
    title_style = ParagraphStyle(
        "title_vo", parent=styles["Title"],
        fontSize=14, textColor=blue, alignment=1, spaceAfter=6,
    )
    elements.append(Paragraph("PROGRAMACAO DE DRAGAGEM", title_style))

    # ---- Info do projeto ----
    info_style = ParagraphStyle(
        "info_vo", parent=styles["Normal"], fontSize=8, spaceAfter=2,
    )
    semana_fim = semana_inicio + timedelta(days=6)
    empresa = config.get("empresa", "")
    vessel = config.get("vessel", "")
    project = config.get("project", "")
    site = config.get("site", "")
    elements.append(Paragraph(
        f"<b>{empresa}</b> &nbsp;&nbsp;|&nbsp;&nbsp; Project: {project}", info_style
    ))
    elements.append(Paragraph(
        f"Vessel: {vessel} &nbsp;&nbsp;|&nbsp;&nbsp; Site: {site}", info_style
    ))
    elements.append(Paragraph(
        f"Semana: {semana_inicio.strftime('%d/%m/%Y')} - {semana_fim.strftime('%d/%m/%Y')}",
        info_style,
    ))
    elements.append(Spacer(1, 4 * mm))

    # ---- Tabela ----
    col_headers = [
        "Data", "Status", "Ampl.", "Inicio", "Final", "Area",
        "KP In.", "KP Fin.", "De", "Ate", "Sist.", "Obs",
        "Dist.", "Tela", "Nozzle", "Dir.", "V.Fr", "V.Re", "Diesel",
    ]

    table_data = [col_headers]

    last_date = ""
    for t in trechos:
        show_date = t["data"] != last_date
        last_date = t["data"]
        table_data.append([
            t["data"] if show_date else "",
            t["status"],
            t["amplitude"],
            t["inicio"],
            t["fim"],
            t.get("area") or "",
            t.get("kp_inicio") or "",
            t.get("kp_final") or "",
            t.get("linha_de") or "",
            t.get("linha_ate") or "",
            t.get("sistema_dragagem") or "",
            t.get("observacoes") or "",
            t.get("dist_fundo") or "",
            t.get("ang_tela") or "",
            t.get("ang_nozzle") or "",
            t.get("direcao") or "",
            t.get("vel_frente") or "",
            t.get("vel_re") or "",
            t.get("consumo_diesel") or "",
        ])

    # Larguras proporcionais ao espaco disponivel (landscape A4 ~= 277mm util)
    avail = landscape(A4)[0] - 20 * mm
    col_pcts = [
        .065, .04, .04, .045, .045, .05, .04, .04, .04, .04,
        .04, .08, .05, .04, .045, .04, .04, .04, .045,
    ]
    col_widths = [avail * p for p in col_pcts]

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Estilo da tabela
    style_cmds = [
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), blue),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D0D0")),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]

    # Linhas alternadas + cores de status
    for i in range(1, len(table_data)):
        # Fundo alternado
        if (i - 1) % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), bg_even))

        # Cor do status (coluna 1)
        status = table_data[i][1]
        if status == "EN":
            style_cmds.append(("TEXTCOLOR", (1, i), (1, i), blue))
        elif status == "VZ":
            style_cmds.append(("TEXTCOLOR", (1, i), (1, i), orange))

        # Data em bold
        if table_data[i][0]:
            style_cmds.append(("FONTNAME", (0, i), (0, i), "Helvetica-Bold"))

    tbl.setStyle(TableStyle(style_cmds))
    elements.append(tbl)

    # ---- Legenda ----
    elements.append(Spacer(1, 4 * mm))
    leg_style = ParagraphStyle(
        "legend_vo", parent=styles["Normal"],
        fontSize=6, textColor=colors.HexColor("#666666"), leading=9,
    )
    elements.append(Paragraph(
        "<b>(x)</b> AE = Agitador Externo; AI = Agitador Interno; "
        "COR = Corrente; CL = Curva Lateral; DS = Desagregador; "
        "EH = Erosion Head ou Mass Flow; JB = Jet Bar; JL = Jato Lateral; "
        "RD = Reduction",
        leg_style,
    ))
    elements.append(Paragraph("<b>(xx)</b> Direction of the bow", leg_style))

    doc.build(elements)
    buf.seek(0)
    return buf
